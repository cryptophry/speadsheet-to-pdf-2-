from flask import Flask, request
from flask.helpers import send_from_directory
from werkzeug.utils import secure_filename
import os, traceback
import pandas as pd
import numpy as np
import openpyxl
import matplotlib as mpl
mpl.use('Agg')
import matplotlib.ticker as mtick
from fpdf import FPDF

site = Flask(__name__)
site.config["UPLOAD_DIR"] = "spreadsheets" # for uploaded spreadsheets
site.config["OUTPUT_DIR"] = "reports" # for downloadable PDF reports
site.config["PLOT_DIR"] = "plots" # for data visualization images

for dir in ["UPLOAD_DIR", "OUTPUT_DIR", "PLOT_DIR"]:
    if not os.path.exists(site.config[dir]):
            os.mkdir(site.config[dir])

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'ods'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def process_spreadsheet(filename):

    def floor_currency(value):
        # Round down to two decimal places
        return np.floor(value*100)/100

    # Wrangle spreadsheet
    spreadsheet = pd.read_excel(filename, sheet_name=["Sales", "Inventory"])
    sale_data = spreadsheet["Sales"]
    item_data = spreadsheet["Inventory"]

    # Join item data to sale data
    sale_data = sale_data.join(item_data.set_index("Code"), on="Item Code", how="inner")

    print(sale_data)

    # Group data
    sales_by_salesperson = sale_data.groupby("Sales Person")
    sales_by_item = sale_data.groupby("Item Name")

    # 1. Which sales person brought in the most revenue?

    # 1.1 Calculate revenue for each sale
    sale_data["Sale Revenue"] = sale_data["Quantity Sold"] * \
        (sale_data["Sale Price"] * (1 - sale_data["Discount"]))
    
    #print(sale_data)
    # 1.2 Round down to 2 decimal places
    sale_data["Sale Revenue"] = sale_data["Sale Revenue"].apply(floor_currency)

    # 1.3 Group sales by sales person and sum sale revenue
    revenue_by_salesperson = sales_by_salesperson["Sale Revenue"].sum().sort_values(
        ascending=False)


    # 2. Which sales person brought in the most profit?

    # 2.1 Calculate profit for each sale
    sale_data["Sale Profit"] = sale_data["Sale Revenue"] - \
        (sale_data["Cost Price"] * sale_data["Quantity Sold"])
    # 2.2 Round down to 2 decimal places
    sale_data["Sale Profit"] = sale_data["Sale Profit"].apply(floor_currency)
    # 2.3 Group sales by sales person and sum sale profit
    profit_by_salesperson = sales_by_salesperson["Sale Profit"].sum().sort_values(
        ascending=False)

    # 3. What was our most discounted item on average?
    average_discounts = sales_by_item["Discount"].mean().sort_values(ascending=False)


    # 4. How much of each item do we have left in stock?

    # 4.1 Sort item_data by name to match sales_by_item
    item_data = item_data.sort_values("Item Name")
    # 4.2 Calculate total sold of each item and add it to item_data as a new column
    item_data = item_data.assign(
        StockSold=sales_by_item["Quantity Sold"].sum().values)
    # 4.3 Add a second new column showing the difference between Stock and StockSold
    item_data["StockLeft"] = item_data["Stock"] - item_data["StockSold"]


    # Display results
    print(revenue_by_salesperson.head())
    print(profit_by_salesperson.head())
    print(average_discounts.head())
    print(item_data[["Item Name", "Stock", "StockSold", "StockLeft"]].head())

    # Create sales pie charts
    def salesperson_pie_chart(df, yaxis, filename):
        explode = np.zeros(df.shape[0])
        explode[0] = 0.1

        pie_profit = df.plot(
            y=yaxis,
            kind="pie",
            explode=explode,
            autopct=lambda value: "${:,.2f}".format(
                floor_currency(value/100 * df.sum())))

        pie_profit.get_figure().gca().set_ylabel("")
        pie_profit.get_figure().tight_layout()
        pie_profit.get_figure().savefig(filename)
        pie_profit.get_figure().clf()


    salesperson_pie_chart(revenue_by_salesperson, "Sale Revenue",
        os.path.join(site.config["PLOT_DIR"], "revenue_by_salesperson.png"))
    salesperson_pie_chart(profit_by_salesperson, "Sale Profit",
        os.path.join(site.config["PLOT_DIR"], "profit_by_salesperson.png"))

    # Create bar chart
    bar_ave_discount = average_discounts.plot(
                        y="Discount",
                        kind="bar",
                        rot=45)

    bar_ave_discount.get_figure().gca().set_xlabel("")
    bar_ave_discount.get_figure().tight_layout()
    bar_ave_discount.yaxis.set_major_formatter(
        mtick.PercentFormatter(xmax=1.0, decimals=0))
    bar_ave_discount.get_figure().savefig(
        os.path.join(site.config["PLOT_DIR"],
            "item_average_discount.png"))
    bar_ave_discount.get_figure().clf()


    # Create PDF
    pdf = FPDF('L') # landscape
    pdf.add_page() # first page
    pdf.set_font('arial', '', 12) # 12pt Arial text


    # Pie charts
    pdf.cell(112, # width
            20, # height
            "Revenue by salesperson", # text
            0, # border (0 = none, 1 = border)
            0, # where to put the cursor for the next cell
                # (0 = right, 1 = next line, 2 = directly below)
            "L") # text alignment

    pdf.cell(110,20, "Profit by salesperson", 0, 1, "L")


    start_x = pdf.get_x()
    start_y = pdf.get_y()

    pdf.image(os.path.join(site.config["PLOT_DIR"],
            "revenue_by_salesperson.png"), w=110)


    pdf.set_xy(start_x + 110 + 2, start_y)
    pdf.image(os.path.join(site.config["PLOT_DIR"],
                "profit_by_salesperson.png"), w=110)

    below_pie_y = pdf.get_y()


    # Stock table
    pdf.set_font('arial', 'B', 10) # table heading font
    pdf.set_y(start_y)
    pdf.set_x(start_x + 220)

    pdf.cell(30, 10, "Item", 1, 0, "C")
    pdf.cell(30, 10, "Stock Left", 1, 2, "C")
    pdf.cell(-30)


    pdf.set_font('arial', '', 10) # table rows font
    for _, row in item_data.iterrows():
        pdf.set_x(start_x + 220)
        pdf.cell(30, 10, row["Item Name"], 1, 0, "L")
        pdf.cell(30, 10, str(row["StockLeft"]), 1, 2, "R")
        pdf.cell(-30)


    # Bar chart
    pdf.set_font('arial', '', 12) # 12pt Arial text
    pdf.set_xy(start_x, below_pie_y-10)
    pdf.cell(30, 10, "Average discounts", 0, 2, "L")
    pdf.image(os.path.join(site.config["PLOT_DIR"],
                "item_average_discount.png"), w=103)

    return pdf.output(os.path.join(site.config["OUTPUT_DIR"],
                    "report.pdf"), "F")

@site.route('/')
def index():
        return """
<!DOCTYPE html>
<html>
  <head>
    <title>Link Results Report Generator</title>
  </head>
  <body>
    <h1>Link Results Report Generator</h1>
    <form action="/process" method="post" enctype="multipart/form-data">
      <input type="file" name="file">
      <input type="submit" value="Generate Report">
    </form>
  </body>
</html>
"""

@site.route('/process', methods=["POST"])
def upload_and_process():
    if "file" not in request.files: # invalid request
        return "Invalid request."

    file = request.files['file']
    if file.filename == '': # no file uploaded by user
            return "No file selected."

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file.save(os.path.join(site.config['UPLOAD_DIR'], filename))

        try:
            process_spreadsheet(os.path.join(site.config['UPLOAD_DIR'],filename))
        except Exception as e:
            print(e)
            traceback.print_exc()
            return "An error occurred. Please ensure that your spreadsheet is correctly formatted and try again."
        else:
            return send_from_directory(directory=site.config["OUTPUT_DIR"], path="report.pdf")

site.run(host='0.0.0.0', port=8080)